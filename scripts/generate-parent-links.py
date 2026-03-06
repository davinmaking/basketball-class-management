#!/usr/bin/env python3
"""Generate Excel spreadsheet with student parent portal links."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_URL = "https://basketball-class-management.vercel.app/view"

students = [
    {"name": "Adam Ting Tin Hao", "class": "Tahun 1", "token": "32eb71c069944185d0d2b6c0111f817a"},
    {"name": "Jared Juk Julian", "class": "Tahun 1", "token": "0d311d3bb1ac99c2fbaab217146e6ba1"},
    {"name": "Lee Kai En", "class": "Tahun 1", "token": "28ec2cc0f4f86c56a4e58bc9ba028c16"},
    {"name": "Muhammad Fikri Hakim", "class": "Tahun 1", "token": "62c5f26c3e008563efbbccadbe7a0947"},
    {"name": "Muhammad Umar Syafiee Bin Matu", "class": "Tahun 1", "token": "e9935aec57a42ec6ab5bfb2ebc8da3e5"},
    {"name": "Siek Tse Zhen", "class": "Tahun 1", "token": "9324228641b28ecb84664d15aa247c9d"},
    {"name": "Alfred Chong Guang Yuan", "class": "Tahun 2", "token": "44dd7123dba68eebd72aa79badb3c57f"},
    {"name": "Ang Yee Xuan", "class": "Tahun 2", "token": "ba0a4e784e3f2980db56529ba4889df2"},
    {"name": "Devynna Hanna Anak Desmond", "class": "Tahun 2", "token": "ac8783799ab0fa8fdab3db06c386a24e"},
    {"name": "Donald Tieng Yew Cheng", "class": "Tahun 2", "token": "f2185c0b316d9b5ab5a338972ff13570"},
    {"name": "Marthinus Leam Menjang", "class": "Tahun 2", "token": "33543be9f5a7bbae13702bb52871e05a"},
    {"name": "Nur Annisa Alliya", "class": "Tahun 2", "token": "79b8b0053d8af5cb4e9334726c3e4e3e"},
    {"name": "Raphael Usat", "class": "Tahun 2", "token": "30b14624c548031a4ecd93ca46b22bc5"},
    {"name": "Rivera Dyesebell Chong Ziyi", "class": "Tahun 2", "token": "eb88994f92d5ab5a65a1b5e56ea68a85"},
    {"name": "Zoey Ng Pey Ling", "class": "Tahun 2", "token": "ea82e98a03434b5cb5b0197a100e5eec"},
    {"name": "Abraham Davinno Elmuda", "class": "Tahun 3", "token": "b5728b586f731a10df8ba8f253dd5712"},
    {"name": "Bong Yi Yang", "class": "Tahun 3", "token": "1422bbee91b438e516af29a17eaa93db"},
    {"name": "Hailey Chin Qian Qian", "class": "Tahun 3", "token": "db38e246de64a070b32215ca1f3e114f"},
    {"name": "Vickie Anak Edit", "class": "Tahun 3", "token": "3ed0e777ab454c42ebab5934dfc59f85"},
    {"name": "Ziv Yu Zi", "class": "Tahun 3", "token": "619e42400d8955f59ec41950c01f0196"},
    {"name": "Celine Leong Jean Wee", "class": "Tahun 4", "token": "dcd38f82c2c1391beb4e68a24adaf70f"},
    {"name": "Charel Cheang Yue Ying", "class": "Tahun 4", "token": "e60fc4ee87f91ef99eb680969a83ff8d"},
    {"name": "Emily Ngu Dai King", "class": "Tahun 4", "token": "c0a090bcfd16a19c930785e596e1b291"},
    {"name": "Jong Huang Dong", "class": "Tahun 4", "token": "6bcea8677c0c45468e18edfa5fe48dc3"},
    {"name": "Lee Jin Fong", "class": "Tahun 4", "token": "2ebe44bd3d69d3255fe6075e6a93d844"},
    {"name": "Megleumim Leah Mayang", "class": "Tahun 4", "token": "3528ef62bcc274a78e748f0e87f13f2e"},
    {"name": "Mohamad Azlan", "class": "Tahun 4", "token": "9a3d7799994bf26d0af34d09c52200f8"},
    {"name": "Terry Chin Zheng Zhe", "class": "Tahun 4", "token": "c1267c9f441d5f2374775adc61320223"},
    {"name": "Chai Wan Ying", "class": "Tahun 5", "token": "27e2a319993703d20cc976cb922795bc"},
    {"name": "Colbie Kupa", "class": "Tahun 5", "token": "eccb9659711b84e6149bc04a9605980e"},
    {"name": "Joice Ngu Dai Ping", "class": "Tahun 5", "token": "319c3fc8639d73b1c34c1d1d27953dc7"},
    {"name": "Jong Nian Ying", "class": "Tahun 5", "token": "6bd60670ca376f7c18c9022cdf832bd2"},
    {"name": "Louis Ng Jun Liang", "class": "Tahun 5", "token": "506416301a074823421b48efd6c22d39"},
    {"name": "Muhammad Fahmie Hakimi", "class": "Tahun 5", "token": "7f5722e1a970a0ec480fc5f179ee5f57"},
    {"name": "Sean Vermont Siaw", "class": "Tahun 5", "token": "64c831a6c7f97a80619be4f07be67c83"},
    {"name": "Abigail Yii Chien Hui", "class": "Tahun 6", "token": "9667ed975aa4762b0c8df01113bc705b"},
    {"name": "Chang Jing Huai", "class": "Tahun 6", "token": "33956e59588fe2d32fae90a44e2c24e0"},
    {"name": "Evander Among Anak Bony", "class": "Tahun 6", "token": "203097732e79209c3c428379933f1177"},
    {"name": "Gloria Anak Wat", "class": "Tahun 6", "token": "02e07acf81ca08cc26621cb9cb2553f0"},
    {"name": "Harraz Bin Jumali", "class": "Tahun 6", "token": "49d246c86eefbfed6935df4abc2f3681"},
    {"name": "Jayden Chin Hao Yu", "class": "Tahun 6", "token": "75dfba6d6065d35410cd477d7a53f4c8"},
    {"name": "Jong Huang Jie", "class": "Tahun 6", "token": "f23869746429d55d79086ed0cf1c1f6b"},
    {"name": "Lee Jing Yii", "class": "Tahun 6", "token": "a886c268b9b21309efece08ca7d46da0"},
    {"name": "Mc Allan", "class": "Tahun 6", "token": "3d1d989ed7d92723623a192219e4142f"},
    {"name": "Muhammad Rahmatdani Shafie Bin Matu", "class": "Tahun 6", "token": "64f694986f54b964de6f113624192dd2"},
    {"name": "Natalia Anak Alexius Gelanggang", "class": "Tahun 6", "token": "809c22e906e0a7ead0288ea922e2b85b"},
    {"name": "Nur Alizsa", "class": "Tahun 6", "token": "b6b3019e6747a97064cce411df3f8fef"},
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "家长链接"

# Styles
header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2E6B30", end_color="2E6B30", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")

class_header_font = Font(name="Arial", bold=True, size=11, color="2E6B30")
class_header_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

data_font = Font(name="Arial", size=11)
link_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
link_fill = PatternFill(start_color="2E6B30", end_color="2E6B30", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Title row
ws.merge_cells("A1:C1")
title_cell = ws["A1"]
title_cell.value = "篮球训练班 - 家长查看链接 / Pautan Portal Ibu Bapa"
title_cell.font = Font(name="Arial", bold=True, size=14, color="2E6B30")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# Subtitle
ws.merge_cells("A2:C2")
subtitle_cell = ws["A2"]
subtitle_cell.value = "请点击链接查看孩子的记录 / Sila klik pautan untuk melihat rekod anak anda"
subtitle_cell.font = Font(name="Arial", size=10, italic=True, color="666666")
subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# Empty row
ws.row_dimensions[3].height = 8

# Headers
headers = ["No.", "学生姓名 / Nama Pelajar", "链接 / Pautan"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[4].height = 30

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 22

# Data rows
row = 5
num = 1
current_class = None

for s in students:
    if s["class"] != current_class:
        current_class = s["class"]
        # Class header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=f"  {current_class}")
        cell.font = class_header_font
        cell.fill = class_header_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row].height = 26
        row += 1

    # Number
    cell_num = ws.cell(row=row, column=1, value=num)
    cell_num.font = data_font
    cell_num.alignment = Alignment(horizontal="center", vertical="center")
    cell_num.border = thin_border

    # Name
    cell_name = ws.cell(row=row, column=2, value=s["name"])
    cell_name.font = data_font
    cell_name.alignment = Alignment(vertical="center")
    cell_name.border = thin_border

    # Link
    url = f"{BASE_URL}/{s['token']}"
    cell_link = ws.cell(row=row, column=3)
    cell_link.value = "点击查看 / Klik Sini"
    cell_link.hyperlink = url
    cell_link.font = link_font
    cell_link.fill = link_fill
    cell_link.alignment = Alignment(horizontal="center", vertical="center")
    cell_link.border = thin_border

    ws.row_dimensions[row].height = 24
    num += 1
    row += 1

# Footer note
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
footer = ws.cell(row=row, column=1, value="* 此链接仅供家长使用 / Pautan ini hanya untuk ibu bapa sahaja")
footer.font = Font(name="Arial", size=9, italic=True, color="999999")
footer.alignment = Alignment(horizontal="center")

# Freeze header
ws.freeze_panes = "A5"

# Save
output_dir = os.path.expanduser("~/Desktop")
output_path = os.path.join(output_dir, "Basketball-Parent-Links.xlsx")
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Total students: {len(students)}")
